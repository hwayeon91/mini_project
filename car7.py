from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import openpyxl
from openpyxl.utils import get_column_letter

import re
import time

#판다스 임포트
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
#엑셀 강제 저장 위한 라이브러리 추가 
import pyautogui
import subprocess

options = Options()
options.add_argument('--start-maximized')
options.add_experimental_option("detach", True)
options.add_experimental_option("excludeSwitches", ['enable-logging'])

time.sleep(2)


driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})
driver.get('https://www.danawa.com/')
time.sleep(2)


#,"kia","genesis","chevorlet","bmw","benz","audi","volvo"
#,307,304,312,362,349,371,459
titlelist = ["월","월 차모델명","월 판매량"]
#자동차 brand 추가될 시 csv 파일 추가
sheet_names = ["hyundai","kia","genesis","chevorlet","bmw","benz","audi","volvo"]
#자동차 brand 추가될 시 추가 
brand = [303,307,304,312,362,349,371,459]
brand_size = len(brand)

# 새 워크시트(마지막 삭제 예정)
wb = openpyxl.Workbook()
# 월별 토탈 데이터 저장 리스트 생성
monthly_totals = []

for i in range(brand_size):
    brand_code = brand[i]
    sell_car_lists = []
    sell_count_lists = []
    sell_tuples = []
    sheet_name = sheet_names[i]
    # Create a new sheet for each brand
    ws_2022 = wb.create_sheet(title=f'{sheet_name}_2022_data')
    ws_2023 = wb.create_sheet(title=f'{sheet_name}_2023_data')

    # 처음 만든 시트 삭제 

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    for year in ['2022','2023']:
        for month_number  in range(1, 13):
            month = str(month_number)
            
            # 시작 열 계산
            start_column = (month_number - 1) * 3 + 1
            # 각각 달에 타이틀을 추가
            for col_index, header in enumerate(titlelist):
                if year == '2022':
                    ws = ws_2022
                else:
                    ws = ws_2023

                ws.cell(row=1, column=start_column + col_index, value=f"{month_number}"+f"{header}")

            # 데이터 수집
            driver.get(f'https://auto.danawa.com/newcar/?Work=record&Tab=Grand&Brand={str(brand_code)}&Month={str(year)}-{str(month_number).zfill(2)}-00&MonthTo=')
            time.sleep(1.5)  # 페이지가 로딩되기를 기다림
            
            #차량 종류 리스트 수집
            elems_titles = driver.find_elements(By.CSS_SELECTOR, '#autodanawa_gridC .title > a')
            elems_nums = driver.find_elements(By.CSS_SELECTOR, '#autodanawa_gridC .num')

            #hidden data 때문에 데이터가 있는경우만 리스트로 뽑아냄 
            sell_car_lists = [elem.text for elem in elems_titles if elem.text.strip()]
            sell_count_numbers = [int(elem.text.replace(',', '')) for elem in elems_nums if elem.text.strip()]


            #리스트 2개를 튜플로 만들어서 엑셀 삽입
            sell_tuples=list(zip(sell_car_lists, sell_count_numbers))
            for car_index, car_tuple in enumerate(sell_tuples):
                for col_index, value in enumerate([f'{year}-{month}'] + list(car_tuple)):
                    if year == '2022':
                        ws = ws_2022
                    else:
                        ws = ws_2023
                    ws.cell(row=car_index + 2,  column=start_column + col_index, value=value)
        max_row = ws.max_row
        max_col = ws.max_column
        # 각 월별 판매량 계산
        for col_index in range(1, max_col + 1, 3):
            month_column = get_column_letter(col_index)
            total_column = get_column_letter(col_index + 2)

            ws.cell(row=max_row + 1, column=col_index, value='월별합계') 
            month_sum_range = f'=SUM({total_column}2:{total_column}{max_row})'
            ws.cell(row=max_row + 1, column=col_index + 2, value=month_sum_range)

        # 연도 토탈 판매량 계산(월별 합계 부분의 열을 합한다)
        start_column_index = 3
        column_index_pattern = 3
        yearly_sum = 0
        # 합산 수식 생성
        sum_range = ",".join([f"{get_column_letter(start_column_index + (column_index_pattern * i))}{ws.max_row}" for i in range(12)])
        yearly_sum = f"=SUM({sum_range})"
        ws.cell(row=max_row + 2, column=1, value='연별합계')
        ws.cell(row=max_row + 2, column=2, value=yearly_sum)
      
wb.save('car.xlsx') 
wb.close()

driver.quit()
# 엑셀을 한번 다시 저장해줘야 데이터를 읽을 수 있으므로 아래를 추가함
pyautogui.hotkey('winleft', 'd')  
time.sleep(1)

# 실행창 Win+R
pyautogui.hotkey('winleft', 'r')  
time.sleep(1)
pyautogui.write('D:\\hwayeon\\workspace\\project\\car.xlsx')  # Type 'excel' and press Enter to open Excel
pyautogui.press('enter')
time.sleep(5)  # Wait for Excel to open (adjust the delay based on your system's speed)
# Save the Excel file with the same name
pyautogui.hotkey('ctrl', 's')
time.sleep(1)  # Wait for the Save dialog to appear

# Press Enter to confirm the save action
pyautogui.press('enter')
time.sleep(1)  # Wait for the file to be saved


# Close the Excel application
subprocess.run(['taskkill', '/f', '/im', 'excel.exe'])

# 그래프를 위한 리스트 만들기 

# Set up lists to store monthly totals for each year and brand
monthly_totals_2022_list = []
monthly_totals_2023_list = []

#한글폰트깨짐
plt.rcParams['font.family'] ='Malgun Gothic'
plt.rcParams['axes.unicode_minus'] =False
wb = openpyxl.load_workbook('car.xlsx', read_only=False, data_only=True)


sheet_size = len(sheet_names)
for i in range(sheet_size):
    monthly_totals_2022 = []
    monthly_totals_2023 = []

    sheet_name = sheet_names[i]
    for year in ['2022','2023']:
        for month_number in range(1, 13):
            sheet = wb[f'{sheet_name}_{year}_data']
            cell_value = sheet.cell(row=sheet.max_row-1, column=3 * month_number).value
            print(cell_value)
            
            if year == '2022':
                monthly_totals_2022.append(cell_value)
            else:
                monthly_totals_2023.append(cell_value)
    # Append the brand's monthly totals to the overall list
    monthly_totals_2022_list.append(monthly_totals_2022)
    monthly_totals_2023_list.append(monthly_totals_2023)

# Print or use the lists as needed
print(monthly_totals_2022_list)
print(monthly_totals_2023_list)

# Set up the x-axis labels (months)
months = [f"{month}" for month in range(1, 13)]

# Plotting for 2022
plt.figure(figsize=(10, 6))
for i, brand in enumerate(sheet_names):
    plt.plot(months, monthly_totals_2022_list[i], marker='o', label=f"{brand} 2022")

# Add labels and title
plt.xlabel('Month')
plt.ylabel('Total Sales')
plt.title('Monthly Total Sales in 2022')
plt.legend()
plt.grid(True)
plt.savefig('Monthly Total Sales in 2022.png')
plt.show()

# Plotting for 2023
plt.figure(figsize=(10, 6))
for i, brand in enumerate(sheet_names):
    plt.plot(months, monthly_totals_2023_list[i], marker='o', label=f"{brand} 2023")

# Add labels and title
plt.xlabel('Month')
plt.ylabel('Total Sales')
plt.title('Monthly Total Sales in 2023')
plt.legend()
plt.grid(True)
plt.savefig('Monthly Total Sales in 2023.png')
plt.show()

# Data for 2022
total_2022 = [sum(monthly_totals) for monthly_totals in monthly_totals_2022_list]

# Data for 2023
total_2023 = [sum(monthly_totals) for monthly_totals in monthly_totals_2023_list]

# Brand names
brands = sheet_names

# Plotting for 2022
plt.figure(figsize=(10, 6))
plt.pie(total_2022, labels=brands, autopct='%1.1f%%', startangle=90)
plt.title('Brand Market Share in 2022')
legend = plt.legend(labels=brands, loc='center left', bbox_to_anchor=(1, 0.5))  # Adding legend to the right of the plot
legend.set_bbox_to_anchor((1.05, 0.5, 0.2, 0.2))
plt.savefig('Brand_Market_Share_2022.png')
plt.show()

# Plotting for 2023
plt.figure(figsize=(10, 6))
plt.pie(total_2023, labels=brands, autopct='%1.1f%%', startangle=90)
plt.title('Brand Market Share in 2023')
legend = plt.legend(labels=brands, loc='center left', bbox_to_anchor=(1, 0.5))  # Adding legend to the right of the plot
legend.set_bbox_to_anchor((1.05, 0.5, 0.2, 0.2))
plt.savefig('Brand_Market_Share_2023.png')
plt.show()


# 관심있는 브랜드 추출
interest_brands = ["genesis", "bmw", "audi", "benz"]
interest_brand_indices = [sheet_names.index(brand) for brand in interest_brands]
# 막대 그래프 컬러 정의
brand_colors = {'genesis': 'red', 'bmw': 'blue', 'audi': 'orange', 'benz': 'green'}

# 2022년 그래프 
plt.figure(figsize=(10, 6))
bar_width = 0.2  # 막대그래프 넓이 지정 
for i, brand in enumerate(interest_brands):
    plt.bar(
        [month_index + i * bar_width for month_index in range(len(months))],
        monthly_totals_2022_list[sheet_names.index(brand)],
        width=bar_width,
        label=f"{brand} 2022",
        color=brand_colors[brand],
        alpha=0.7
    )

# 라벨과 
plt.xlabel('Month')
plt.ylabel('Total Sales')
plt.title('Monthly Total Sales for Specific Brands in 2022')
plt.xticks([month_index + bar_width for month_index in range(len(months))], months)
plt.legend()
plt.grid(True)
plt.savefig('Monthly Total Sales Specific Brands 2022 Grouped Bar.png')
plt.show()

# Plotting for 2023 (Grouped Bar Graph)
plt.figure(figsize=(10, 6))
for i, brand in enumerate(interest_brands):
    plt.bar(
        [month_index + i * bar_width for month_index in range(len(months))],
        monthly_totals_2023_list[sheet_names.index(brand)],
        width=bar_width,
        label=f"{brand} 2023",
        color=brand_colors[brand],
        alpha=0.7
    )

# Add labels and title
plt.xlabel('Month')
plt.ylabel('Total Sales')
plt.title('Monthly Total Sales for Specific Brands in 2023')
plt.xticks([month_index + bar_width for month_index in range(len(months))], months)
plt.legend()
plt.grid(True)
plt.savefig('Monthly Total Sales Specific Brands 2023 Grouped Bar.png')
plt.show()

# 파이 차트 그리기
total_2022_interest = [total_2022[i] for i in interest_brand_indices]

# Plotting for 2022 (Pie Chart)
plt.figure(figsize=(10, 6))
plt.pie(total_2022_interest, labels=interest_brands, autopct='%1.1f%%', startangle=90)
plt.title('Market Share for Specific Brands in 2022')
legend = plt.legend(labels=interest_brands, loc='center left', bbox_to_anchor=(1, 0.5))
legend.set_bbox_to_anchor((1.05, 0.5, 0.2, 0.2))
plt.savefig('Market Share Specific Brands 2022.png')
plt.show()

# Data for 2023 (Pie Chart)
total_2023_interest = [total_2023[i] for i in interest_brand_indices]

# Plotting for 2023 (Pie Chart)
plt.figure(figsize=(10, 6))
plt.pie(total_2023_interest, labels=interest_brands, autopct='%1.1f%%', startangle=90)
plt.title('Market Share for Specific Brands in 2023')
legend = plt.legend(labels=interest_brands, loc='center left', bbox_to_anchor=(1, 0.5))
legend.set_bbox_to_anchor((1.05, 0.5, 0.2, 0.2))
plt.savefig('Market Share Specific Brands 2023.png')
plt.show()

