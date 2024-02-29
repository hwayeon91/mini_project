from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import time

import cx_Oracle

options = Options()
options.add_argument('--start-maximized')
options.add_experimental_option("detach", True)
options.add_experimental_option("excludeSwitches", ['enable-logging'])

driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})

fpath = r'D:\hwayeon\project\초등교과.xlsx'
workbook = openpyxl.load_workbook(fpath)

# subject = ["국어", "영어", "수학", "사회", "과학", "통합교과(바슬즐)", "예체능", "전과목"]
# grade = ["1학년", "2학년", "3학년", "4학년", "5학년", "6학년"]
# code_number_1_grade = [39010101, 39010102, 39010107, 39010115, 39010117] # 국어, 통합교과(바슬즐),수학, 예체능, 전과목 순서
# code_number_2_grade = [39010301, 39010302, 39010307, 39010315, 39010317] # 국어, 통합교과(바슬즐),수학, 예체능, 전과목 순서
# code_number_3_grade = [39010501, 39010507, 39010509, 39010511, 39010515] # 국어, 수학, 사회, 과학, 영어 순서
# code_number_4_grade = [39010701, 39010707, 39010709, 39010711, 39010715] # 국어, 수학, 사회, 과학, 영어 순서
# code_number_5_grade = [39010901, 39010907, 39010909, 39010911, 39010915] # 국어, 수학, 사회, 과학, 영어 순서
# code_number_6_grade = [39011101, 39011107, 39011109, 39011111, 39011115] # 국어, 수학, 사회, 과학, 영어 순서

#code_number_grade = code_number_1_grade + code_number_2_grade + code_number_3_grade + code_number_4_grade + code_number_5_grade + code_number_6_grade

category_info = {
    39010101 : ("1학년","국어"),
    39010102 : ("1학년","통합교과(바슬즐)"),
    39010107 : ("1학년","수학"),
    39010115 : ("1학년","예체능"),
    39010117 : ("1학년","전과목"),

    39010301 : ("2학년","국어"),
    39010302 : ("2학년","통합교과(바슬즐)"),
    39010307 : ("2학년","수학"),
    39010315 : ("2학년","예체능"),
    39010317 : ("2학년","전과목"),

    39010501 : ("3학년","국어"),
    39010507 : ("3학년","수학"),
    39010509 : ("3학년","사회"),
    39010511 : ("3학년","과학"),
    39010515 : ("3학년","영어"),

    39010701 : ("4학년","국어"),
    39010707 : ("4학년","수학"),
    39010709 : ("4학년","사회"),
    39010711 : ("4학년","과학"),
    39010715 : ("4학년","영어"),

    39010901 : ("5학년","국어"),
    39010907 : ("5학년","수학"),
    39010909 : ("5학년","사회"),
    39010911 : ("5학년","과학"),
    39010915 : ("5학년","영어"),

    39011101 : ("6학년","국어"),
    39011107 : ("6학년","수학"),
    39011109 : ("6학년","사회"),
    39011111 : ("6학년","과학"),
    39011115 : ("6학년","영어")
}

sheet_name = '초등'
if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
sheet = workbook[sheet_name]

sheet['A1'] = "상품 ID"
sheet['B1'] = "제목"
sheet['C1'] = "저자"
sheet['D1'] = "출판사"
sheet['E1'] = "출판일"
sheet['F1'] = "할인율(%)"
sheet['G1'] = "할인가격(원)"
sheet['H1'] = "정가(원)"
sheet['I1'] = "소개"
sheet['J1'] = "이미지"
sheet['K1'] = "카테고리"
sheet['L1'] = "카테고리"
sheet['M1'] = "카테고리"

if 'Sheet1' in workbook.sheetnames:
    workbook.remove(workbook['Sheet1'])

start_row = 2
for category_num in category_info.keys():
    grade, subject = category_info[category_num]
    for page_num in [1,2]:  # 페이지 번호를 1에서 2까지 반복
        url = f'https://product.kyobobook.co.kr/category/KOR/{category_num}#?page={page_num}&type=all&per=20&sort=sel'
        driver.get(url)
        time.sleep(2)

        # 스크롤 다운 (페이지의 끝까지 스크롤)
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)  # 스크롤 동안 대기
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # 스크래핑
        book_id = driver.find_elements(By.CLASS_NAME, 'prod_item')
        book_title = driver.find_elements(By.CLASS_NAME, 'prod_name')
        book_author = driver.find_elements(By.CLASS_NAME, 'prod_author')
        book_discounted_rate = driver.find_elements(By.CSS_SELECTOR, '.prod_price .percent')
        book_discounted_price = driver.find_elements(By.CLASS_NAME, 'price')
        book_normal_price = driver.find_elements(By.CSS_SELECTOR, '.price_normal .val')
        book_image = driver.find_elements(By.CSS_SELECTOR, '.prod_item .img_box')
        book_introduction = driver.find_elements(By.CLASS_NAME, 'prod_introduction')

        for id, title, author, discounted_rate, discounted_price, normal_price, image, introduction in zip(book_id,
                                                                                                            book_title,
                                                                                                            book_author,
                                                                                                            book_discounted_rate,
                                                                                                            book_discounted_price,
                                                                                                            book_normal_price,
                                                                                                            book_image,
                                                                                                            book_introduction):
            id_text = id.get_attribute('data-id')
            title_text = title.text
            # prod_author '이미례', '리틀씨앤톡', '2024.02.15' 데이터구조
            author_text = author.text.split(" · ")[0]
            publisher_text = author.text.split(" · ")[1]
            publish_date_text = author.text.split(" · ")[2]
            discounted_rate_int = (int)(discounted_rate.text.strip('%'))
            price_text = discounted_price.text
            normal_price_text = normal_price.text
            introduction_text = introduction.text
            image_url = image.find_element(By.TAG_NAME, 'img').get_attribute('src') if book_image else "이미지 URL 없음"

            print("상품ID : ", id_text)
            print("제목 : ", title_text)
            print("저자 : ", author_text)
            print("출판사 : ", publisher_text)
            print("출판일 : ", publish_date_text)
            print("할인율 : ", discounted_rate_int)
            print("할인가격 : ", price_text)
            print("정가 ", normal_price_text)
            print("소개 : ", introduction_text)
            print("이미지 : ", image_url)
            print("카테고리 : ", category_num )
            print("학년 : ", grade )
            print("과목 : ", subject )
            print("\n")
            #product_id 를 스크래핑된 id 로 하게 되면 하나의 책이 두군데 이상의 카테고리 중복되고있어 primary key 로 쓸 수 없음
            #product_id = str(category_num) + id_text
            #sheet[f'A{start_row}'] = product_id
            sheet[f'A{start_row}'] = id_text
            sheet[f'B{start_row}'] = title_text
            sheet[f'C{start_row}'] = author_text
            sheet[f'D{start_row}'] = publisher_text
            sheet[f'E{start_row}'] = publish_date_text
            sheet[f'F{start_row}'] = discounted_rate_int
            sheet[f'G{start_row}'] = price_text
            sheet[f'H{start_row}'] = normal_price_text
            sheet[f'I{start_row}'] = introduction_text
            sheet[f'J{start_row}'] = image_url
            sheet[f'K{start_row}'] = category_num
            sheet[f'L{start_row}'] = grade
            sheet[f'M{start_row}'] = subject
            
            start_row += 1

workbook.save(fpath)

driver.quit()


#DB 접속

con = cx_Oracle.connect("HWAYEON", "HWAYEON", "192.168.0.122:1521/xe", encoding="UTF-8") #오라클 연결
print( con ) #연결확인

cursor = con.cursor() #CRUD명령 실행을 위한 커서 객체를 얻는다.
#쿼리문
try : 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):

        sql_insert = """
            INSERT INTO PRODUCTS (bookId, productId, productTitle, author, publisher, publishDate, 
                                        discountedRate, price, normalPrice, introduction, 
                                        imageUrl, categoryNum, grade, subject, bookStock, regDate, updateDate)
            VALUES (SEQ_BOOKID.NEXTVAL, :1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11, :12, :13, 5, sysdate, sysdate)
        """

        # Execute the INSERT statement
        cursor.execute(sql_insert, row)

    # Commit the changes to the database
    con.commit()

    cursor.execute(sql_insert)
    # x = cursor.fetchall()

    # print("=====>", x)
    con.commit() #커밋을 통한 트랜잭션 종료
except cx_Oracle.IntegrityError as e:
    print(f"IntegrityError: {e}")
finally :
    cursor.close()
    con.close()