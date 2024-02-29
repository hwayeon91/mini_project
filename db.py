#DB 접속
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import time

import cx_Oracle

fpath = r'D:\hwayeon\project\초등교과.xlsx'
workbook = openpyxl.load_workbook(fpath)

sheet_name = '초등'
if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
sheet = workbook[sheet_name]

con = cx_Oracle.connect("HWAYEON", "HWAYEON", "192.168.0.122:1521/xe", encoding="UTF-8") #오라클 연결
print( con ) #연결확인

cursor = con.cursor() #CRUD명령 실행을 위한 커서 객체를 얻는다.
#쿼리문
try : 
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):

        sql_insert = """
            INSERT INTO PRODUCTS (bookId, productId, productTitle, author, publisher, publishDate, 
                                        discountedRate, price, normalPrice, introduction, 
                                        imageUrl, categoryNum, grade, subject, bookStock, regDate, updateDate)
            VALUES (SEQ_BOOKID.NEXTVAL, :1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11, :12, :13, 5, sysdate, sysdate)
        """

        # Execute the INSERT statement
        cursor.execute(sql_insert, row)

    # Commit the changes to the database
    con.commit()

    cursor.execute(sql_insert)
    # x = cursor.fetchall()

    # print("=====>", x)
    con.commit() #커밋을 통한 트랜잭션 종료
except cx_Oracle.IntegrityError as e:
    print(f"IntegrityError: {e}")
finally :
    cursor.close()
    con.close()