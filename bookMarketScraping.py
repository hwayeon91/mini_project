from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import time


options = Options()
options.add_argument('--start-maximized')
options.add_experimental_option("detach", True)
options.add_experimental_option("excludeSwitches", ['enable-logging'])

driver = webdriver.Chrome(options=options)
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})

fpath = r'D:\YDH\workspace\python\프로젝트\초등교과.xlsx'
workbook = openpyxl.load_workbook(fpath)

subject = ["국어", "영어", "수학", "사회", "과학", "통합교과(바슬즐)", "예체능", "전과목"]
grade = ["1학년", "2학년", "3학년", "4학년", "5학년", "6학년"]
code_number_1_grade = [39010101, 39010102, 39010107, 39010115, 39010317] # 국어, 통합교과(바슬즐),수학, 예체능, 전과목 순서
code_number_2_grade = [39010301, 39010302, 39010307, 39030115, 39010317] # 국어, 통합교과(바슬즐),수학, 예체능, 전과목 순서
code_number_3_grade = [39010501, 39010507, 39010509, 39010511, 39010515] # 국어, 수학, 사회, 과학, 영어 순서
code_number_4_grade = [39010701, 39010707, 39010709, 39010711, 39010715] # 국어, 수학, 사회, 과학, 영어 순서
code_number_5_grade = [39010901, 39010907, 39010909, 39010911, 39010915] # 국어, 수학, 사회, 과학, 영어 순서
code_number_6_grade = [39011101, 39011107, 39011109, 39011111, 39011115] # 국어, 수학, 사회, 과학, 영어 순서

sheet_name = '초등'
if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
sheet = workbook[sheet_name]

sheet['A1'] = "상품 ID"
sheet['B1'] = "제목"
sheet['C1'] = "저자"
sheet['D1'] = "출판사"
sheet['E1'] = "출판일"
sheet['F1'] = "할인율"
sheet['G1'] = "할인가격"
sheet['H1'] = "정가"
sheet['I1'] = "소개"
sheet['J1'] = "이미지"

start_row = 2

for page_num in range(1, 20):  # 페이지 번호를 1에서 2까지 반복
    url = f'https://product.kyobobook.co.kr/category/KOR/39010101#?page={page_num}&type=all&sort=sel'
    driver.get(url)
    time.sleep(2)

    # 스크롤 다운 (페이지의 끝까지 스크롤)
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)  # 스크롤 동안 대기
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # 스크래핑
    book_id = driver.find_elements(By.CLASS_NAME, 'prod_item')
    book_title = driver.find_elements(By.CLASS_NAME, 'prod_name')
    book_author = driver.find_elements(By.CLASS_NAME, 'prod_author')
    book_discounted_rate = driver.find_elements(By.CSS_SELECTOR, '.prod_price .percent')
    book_discounted_price = driver.find_elements(By.CLASS_NAME, 'price')
    book_normal_price = driver.find_elements(By.CSS_SELECTOR, '.price_normal .val')
    book_image = driver.find_elements(By.CSS_SELECTOR, '.prod_item .img_box')
    book_introduction = driver.find_elements(By.CLASS_NAME, 'prod_introduction')

    for id, title, author, discounted_rate, discounted_price, normal_price, image, introduction in zip(book_id,
                                                                                                        book_title,
                                                                                                        book_author,
                                                                                                        book_discounted_rate,
                                                                                                        book_discounted_price,
                                                                                                        book_normal_price,
                                                                                                        book_image,
                                                                                                        book_introduction):
        id_text = id.get_attribute('data-id')
        title_text = title.text
        # prod_author '이미례', '리틀씨앤톡', '2024.02.15' 데이터구조
        author_text = author.text.split(" · ")[0]
        publisher_text = author.text.split(" · ")[1]
        publish_date_text = author.text.split(" · ")[2]
        discounted_rate_text = discounted_rate.text
        price_text = discounted_price.text
        normal_price_text = normal_price.text
        introduction_text = introduction.text
        image_url = image.find_element(By.TAG_NAME, 'img').get_attribute('src') if book_image else "이미지 URL 없음"
        print("상품ID : ", id_text)
        print("제목 : ", title_text)
        print("저자 : ", author_text)
        print("출판사 : ", publisher_text)
        print("출판일 : ", publish_date_text)
        print("할인율 : ", discounted_rate_text)
        print("할인가격 : ", price_text)
        print("정가 ", normal_price_text)
        print("소개 : ", introduction_text)
        print("이미지 : ", image_url)
        print("\n")
        sheet[f'A{start_row}'] = id_text
        sheet[f'B{start_row}'] = title_text
        sheet[f'C{start_row}'] = author_text
        sheet[f'D{start_row}'] = publisher_text
        sheet[f'E{start_row}'] = publish_date_text
        sheet[f'F{start_row}'] = discounted_rate_text
        sheet[f'G{start_row}'] = price_text
        sheet[f'H{start_row}'] = normal_price_text
        sheet[f'I{start_row}'] = introduction_text
        sheet[f'J{start_row}'] = image_url
        start_row += 1

workbook.save(fpath)
driver.quit()
